#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
batch_mannwhitney.py

반복 Mann-Whitney U test script.

기본 input columns:
- step_seq
- item_id
- value
- 거리 구분
- Group

분석 로직:
1. Group이 빈칸이면 "EE-type"으로 변환한다.
2. 같은 step_seq, item_id, Group 안에서만 비교한다.
3. 거리 구분이 E4인 그룹을 reference로 둔다.
4. 같은 key 안의 나머지 거리 구분 그룹들을 각각 E4와 1:1 비교한다.
5. 각 1:1 비교마다 ref와 comp 각각에 대해 IQR rule-of-thumb outlier 제거를 수행한다.
   - lower = Q1 - 1.5 * IQR
   - upper = Q3 + 1.5 * IQR
6. median은 outlier 제거 전 raw data 기준으로 계산한다.
7. p-value는 outlier 제거 후 Mann-Whitney U test로 계산한다.
8. 결과는 Excel로 저장한다.

사용 예시:
    python batch_mannwhitney.py -i input.xlsx -o mannwhitney_result.xlsx

필요 패키지:
    pip install pandas numpy scipy openpyxl
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
from scipy.stats import mannwhitneyu


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Run repeated Mann-Whitney U tests using E4 as reference."
    )

    parser.add_argument(
        "-i",
        "--input",
        required=True,
        help="Input file path. Supported: .xlsx, .xls, .xlsm, .csv, .tsv",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="mannwhitney_result.xlsx",
        help="Output file path. Recommended: .xlsx",
    )

    parser.add_argument("--step-col", default="step_seq", help="Step column name.")
    parser.add_argument("--item-col", default="item_id", help="Item column name.")
    parser.add_argument("--value-col", default="value", help="Numeric value column name.")
    parser.add_argument("--dist-col", default="거리 구분", help="Distance group column name.")
    parser.add_argument("--group-col", default="Group", help="Group column name.")

    parser.add_argument(
        "--ref-dist",
        default="E4",
        help='Reference distance group. Default: "E4"',
    )
    parser.add_argument(
        "--blank-group-name",
        default="EE-type",
        help='Name used when Group is blank. Default: "EE-type"',
    )
    parser.add_argument(
        "--min-n-outlier",
        type=int,
        default=4,
        help="Minimum n required to apply IQR outlier removal. Default: 4",
    )
    parser.add_argument(
        "--case-sensitive-dist",
        action="store_true",
        help=(
            "Use case-sensitive matching for distance groups. "
            "By default, distance labels are uppercased."
        ),
    )
    parser.add_argument(
        "--no-fdr",
        action="store_true",
        help="Do not calculate Benjamini-Hochberg FDR-adjusted p-values.",
    )

    return parser.parse_args()


def read_input_table(input_path: Path) -> pd.DataFrame:
    suffix = input_path.suffix.lower()

    if suffix in [".xlsx", ".xls", ".xlsm"]:
        df = pd.read_excel(input_path)

    elif suffix in [".csv", ".tsv"]:
        sep = "\t" if suffix == ".tsv" else ","
        last_error = None

        for encoding in ["utf-8-sig", "utf-8", "cp949", "euc-kr"]:
            try:
                df = pd.read_csv(input_path, sep=sep, encoding=encoding)
                break
            except UnicodeDecodeError as exc:
                last_error = exc
        else:
            raise UnicodeDecodeError(
                "unknown",
                b"",
                0,
                1,
                (
                    "Could not read CSV/TSV with utf-8-sig, utf-8, cp949, "
                    f"or euc-kr. Last error: {last_error}"
                ),
            )

    else:
        raise ValueError(
            f"Unsupported input file extension: {suffix}. "
            "Use .xlsx, .xls, .xlsm, .csv, or .tsv."
        )

    df.columns = [str(col).strip() for col in df.columns]
    return df


def validate_columns(
    df: pd.DataFrame,
    step_col: str,
    item_col: str,
    value_col: str,
    dist_col: str,
    group_col: str,
) -> None:
    required_cols = [step_col, item_col, value_col, dist_col, group_col]
    missing_cols = [col for col in required_cols if col not in df.columns]

    if missing_cols:
        available = ", ".join(map(str, df.columns))
        missing = ", ".join(missing_cols)
        raise ValueError(
            f"Missing required column(s): {missing}\n"
            f"Available columns: {available}"
        )


def clean_group_name(value, blank_group_name: str) -> str:
    """
    Group column cleaning.
    빈칸, NaN, None, null 등은 EE-type으로 처리한다.
    """
    if pd.isna(value):
        return blank_group_name

    text = str(value).strip()

    if text == "" or text.lower() in ["nan", "none", "null"]:
        return blank_group_name

    return text


def clean_distance_label(value, case_sensitive: bool = False):
    """
    거리 구분 cleaning.
    기본적으로 대소문자 차이를 없애기 위해 upper-case 변환한다.
    """
    if pd.isna(value):
        return np.nan

    text = str(value).strip()

    if text == "" or text.lower() in ["nan", "none", "null"]:
        return np.nan

    return text if case_sensitive else text.upper()


def coerce_numeric(series: pd.Series) -> pd.Series:
    """
    value column을 numeric으로 변환한다.
    예: "1,234.5" -> 1234.5
    """
    if pd.api.types.is_numeric_dtype(series):
        return series.astype(float)

    cleaned = (
        series.astype(str)
        .str.strip()
        .str.replace(",", "", regex=False)
        .replace({"": np.nan, "nan": np.nan, "None": np.nan, "NULL": np.nan})
    )

    return pd.to_numeric(cleaned, errors="coerce")


def natural_sort_key(value) -> List:
    """
    E1, E2, E10 같은 값을 자연스럽게 정렬하기 위한 key.
    """
    return [
        int(part) if part.isdigit() else part
        for part in re.split(r"(\d+)", str(value))
    ]


def median_or_nan(values: pd.Series) -> float:
    values = pd.to_numeric(values, errors="coerce").dropna()
    if len(values) == 0:
        return np.nan
    return float(values.median())


def remove_outliers_iqr(
    values: pd.Series,
    min_n_outlier: int = 4,
) -> Tuple[pd.Series, int, float, float, float, float, float]:
    """
    Tukey IQR rule로 outlier 제거.

    반환값:
        clean_values,
        outlier_count,
        q1,
        q3,
        iqr,
        lower_bound,
        upper_bound
    """
    x = pd.to_numeric(pd.Series(values), errors="coerce").dropna().astype(float)

    if len(x) < min_n_outlier:
        return x, 0, np.nan, np.nan, np.nan, np.nan, np.nan

    q1 = float(x.quantile(0.25))
    q3 = float(x.quantile(0.75))
    iqr = q3 - q1

    lower_bound = q1 - 1.5 * iqr
    upper_bound = q3 + 1.5 * iqr

    keep_mask = x.between(lower_bound, upper_bound, inclusive="both")
    clean_x = x.loc[keep_mask]
    outlier_count = int((~keep_mask).sum())

    return (
        clean_x,
        outlier_count,
        q1,
        q3,
        float(iqr),
        float(lower_bound),
        float(upper_bound),
    )


def mannwhitney_test(
    ref_values: pd.Series,
    comp_values: pd.Series,
) -> Tuple[float, float, str]:
    """
    Two-sided Mann-Whitney U test.

    반환값:
        U statistic,
        p-value,
        status
    """
    ref = pd.to_numeric(pd.Series(ref_values), errors="coerce").dropna().astype(float)
    comp = pd.to_numeric(pd.Series(comp_values), errors="coerce").dropna().astype(float)

    if len(ref) == 0 or len(comp) == 0:
        return np.nan, np.nan, "not_tested_empty_after_outlier_removal"

    try:
        result = mannwhitneyu(
            ref.to_numpy(),
            comp.to_numpy(),
            alternative="two-sided",
            method="auto",
        )
        return float(result.statistic), float(result.pvalue), "ok"

    except TypeError:
        # 구버전 scipy에서 method="auto"를 지원하지 않는 경우
        result = mannwhitneyu(
            ref.to_numpy(),
            comp.to_numpy(),
            alternative="two-sided",
        )
        return float(result.statistic), float(result.pvalue), "ok"

    except Exception as exc:
        return np.nan, np.nan, f"test_error: {exc}"


def bh_fdr_correction(p_values: pd.Series) -> np.ndarray:
    """
    Benjamini-Hochberg FDR correction.

    NaN p-value는 NaN으로 유지한다.
    """
    p = pd.to_numeric(pd.Series(p_values), errors="coerce").to_numpy(dtype=float)
    adjusted = np.full(len(p), np.nan)

    valid_mask = ~np.isnan(p)
    p_valid = p[valid_mask]

    if len(p_valid) == 0:
        return adjusted

    order = np.argsort(p_valid)
    ranked = p_valid[order]
    n = len(ranked)

    adjusted_ranked = ranked * n / np.arange(1, n + 1)
    adjusted_ranked = np.minimum.accumulate(adjusted_ranked[::-1])[::-1]
    adjusted_ranked = np.clip(adjusted_ranked, 0, 1)

    adjusted_valid = np.empty(n)
    adjusted_valid[order] = adjusted_ranked

    adjusted[valid_mask] = adjusted_valid
    return adjusted


def build_result_tables(
    df: pd.DataFrame,
    step_col: str,
    item_col: str,
    value_col: str,
    dist_col: str,
    group_col: str,
    ref_dist: str,
    blank_group_name: str,
    min_n_outlier: int,
    case_sensitive_dist: bool,
    calculate_fdr: bool,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    validate_columns(
        df=df,
        step_col=step_col,
        item_col=item_col,
        value_col=value_col,
        dist_col=dist_col,
        group_col=group_col,
    )

    working = df.copy()

    working[group_col] = working[group_col].apply(
        lambda x: clean_group_name(x, blank_group_name=blank_group_name)
    )

    working["_distance_group_clean"] = working[dist_col].apply(
        lambda x: clean_distance_label(x, case_sensitive=case_sensitive_dist)
    )

    working[value_col] = coerce_numeric(working[value_col])

    ref_dist_clean = clean_distance_label(ref_dist, case_sensitive=case_sensitive_dist)

    total_rows_before = len(working)

    # value가 numeric으로 변환되지 않았거나 거리 구분이 비어 있는 row는 분석에서 제외
    working = working.dropna(subset=[value_col, "_distance_group_clean"])

    total_rows_after = len(working)

    result_rows: List[Dict] = []
    skipped_rows: List[Dict] = []

    for (step_value, item_value, group_value), sub in working.groupby(
        [step_col, item_col, group_col],
        dropna=False,
        sort=True,
    ):
        available_distances = sorted(
            [x for x in sub["_distance_group_clean"].dropna().unique()],
            key=natural_sort_key,
        )

        ref_raw = sub.loc[
            sub["_distance_group_clean"] == ref_dist_clean,
            value_col,
        ].dropna()

        comp_distances = [
            dist for dist in available_distances if dist != ref_dist_clean
        ]

        # E4 reference가 없는 key는 skip
        if len(ref_raw) == 0:
            skipped_rows.append(
                {
                    step_col: step_value,
                    item_col: item_value,
                    group_col: group_value,
                    "reason": f"reference_group_not_found: {ref_dist_clean}",
                    "available_distance_groups": ", ".join(map(str, available_distances)),
                    "n_rows_in_key": len(sub),
                }
            )
            continue

        # 비교 대상 group이 없는 key도 skip
        if len(comp_distances) == 0:
            skipped_rows.append(
                {
                    step_col: step_value,
                    item_col: item_value,
                    group_col: group_value,
                    "reason": "no_compare_group_found",
                    "available_distance_groups": ", ".join(map(str, available_distances)),
                    "n_rows_in_key": len(sub),
                }
            )
            continue

        for comp_dist in comp_distances:
            comp_raw = sub.loc[
                sub["_distance_group_clean"] == comp_dist,
                value_col,
            ].dropna()

            # 매 1:1 비교마다 ref와 comp 각각 outlier 제거
            ref_clean, ref_outlier_n, ref_q1, ref_q3, ref_iqr, ref_lower, ref_upper = (
                remove_outliers_iqr(ref_raw, min_n_outlier=min_n_outlier)
            )

            comp_clean, comp_outlier_n, comp_q1, comp_q3, comp_iqr, comp_lower, comp_upper = (
                remove_outliers_iqr(comp_raw, min_n_outlier=min_n_outlier)
            )

            # p-value는 outlier 제거 후 데이터 기준
            u_stat, p_value, test_status = mannwhitney_test(ref_clean, comp_clean)

            result_rows.append(
                {
                    step_col: step_value,
                    item_col: item_value,
                    group_col: group_value,
                    "ref_group": ref_dist_clean,
                    "comp_group": comp_dist,

                    # 요청 median: outlier 제거 전 raw median
                    "ref_median": median_or_nan(ref_raw),
                    "comp_median": median_or_nan(comp_raw),

                    # p-value: outlier 제거 후 Mann-Whitney U test
                    "p_value": p_value,
                    "U_statistic": u_stat,
                    "test_status": test_status,

                    # 검토용 n / outlier 정보
                    "ref_n_raw": int(len(ref_raw)),
                    "comp_n_raw": int(len(comp_raw)),
                    "ref_n_after_outlier_rm": int(len(ref_clean)),
                    "comp_n_after_outlier_rm": int(len(comp_clean)),
                    "ref_outlier_n": int(ref_outlier_n),
                    "comp_outlier_n": int(comp_outlier_n),

                    # 검토용 outlier 제거 후 median
                    "ref_median_after_outlier_rm": median_or_nan(ref_clean),
                    "comp_median_after_outlier_rm": median_or_nan(comp_clean),

                    # 검토용 outlier 기준값
                    "ref_q1_for_outlier": ref_q1,
                    "ref_q3_for_outlier": ref_q3,
                    "ref_iqr_for_outlier": ref_iqr,
                    "ref_outlier_lower_bound": ref_lower,
                    "ref_outlier_upper_bound": ref_upper,

                    "comp_q1_for_outlier": comp_q1,
                    "comp_q3_for_outlier": comp_q3,
                    "comp_iqr_for_outlier": comp_iqr,
                    "comp_outlier_lower_bound": comp_lower,
                    "comp_outlier_upper_bound": comp_upper,

                    "rows_before_numeric_distance_drop": int(total_rows_before),
                    "rows_after_numeric_distance_drop": int(total_rows_after),
                }
            )

    full_columns = [
        step_col,
        item_col,
        group_col,
        "ref_group",
        "comp_group",
        "ref_median",
        "comp_median",
        "p_value",
        "U_statistic",
        "test_status",
        "ref_n_raw",
        "comp_n_raw",
        "ref_n_after_outlier_rm",
        "comp_n_after_outlier_rm",
        "ref_outlier_n",
        "comp_outlier_n",
        "ref_median_after_outlier_rm",
        "comp_median_after_outlier_rm",
        "ref_q1_for_outlier",
        "ref_q3_for_outlier",
        "ref_iqr_for_outlier",
        "ref_outlier_lower_bound",
        "ref_outlier_upper_bound",
        "comp_q1_for_outlier",
        "comp_q3_for_outlier",
        "comp_iqr_for_outlier",
        "comp_outlier_lower_bound",
        "comp_outlier_upper_bound",
        "rows_before_numeric_distance_drop",
        "rows_after_numeric_distance_drop",
    ]

    result = pd.DataFrame(result_rows, columns=full_columns)

    # 반복 검정이 많으므로 FDR 보정 p-value도 full_result에 추가
    if calculate_fdr and len(result) > 0:
        result.insert(
            result.columns.get_loc("U_statistic"),
            "p_adj_BH_FDR_global",
            bh_fdr_correction(result["p_value"]),
        )

        result.insert(
            result.columns.get_loc("U_statistic"),
            "p_adj_BH_FDR_within_step_item_group",
            np.nan,
        )

        for _, idx in result.groupby(
            [step_col, item_col, group_col],
            dropna=False,
        ).groups.items():
            result.loc[idx, "p_adj_BH_FDR_within_step_item_group"] = bh_fdr_correction(
                result.loc[idx, "p_value"]
            )

    skipped_columns = [
        step_col,
        item_col,
        group_col,
        "reason",
        "available_distance_groups",
        "n_rows_in_key",
    ]

    skipped = pd.DataFrame(skipped_rows, columns=skipped_columns)

    return result, skipped


def build_method_note(
    step_col: str,
    item_col: str,
    value_col: str,
    dist_col: str,
    group_col: str,
    ref_dist: str,
    blank_group_name: str,
    min_n_outlier: int,
    case_sensitive_dist: bool,
    calculate_fdr: bool,
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "item": "Grouping key",
                "detail": f"{step_col}, {item_col}, {group_col}",
            },
            {
                "item": "Reference group",
                "detail": ref_dist,
            },
            {
                "item": "Comparison logic",
                "detail": (
                    "Within each grouping key, every non-reference distance group "
                    "is compared one-by-one against the reference group."
                ),
            },
            {
                "item": "Blank Group handling",
                "detail": (
                    f"Blank or missing {group_col} values are replaced with "
                    f"{blank_group_name}."
                ),
            },
            {
                "item": "Outlier removal",
                "detail": (
                    "For every ref-vs-comp comparison, outliers are removed from "
                    "ref and comp separately using Tukey IQR rule: "
                    "Q1 - 1.5*IQR <= x <= Q3 + 1.5*IQR."
                ),
            },
            {
                "item": "Minimum n for outlier removal",
                "detail": f"IQR outlier removal is applied only when n >= {min_n_outlier}.",
            },
            {
                "item": "Median columns",
                "detail": "ref_median and comp_median are calculated before outlier removal.",
            },
            {
                "item": "p-value",
                "detail": (
                    "p_value is calculated by two-sided Mann-Whitney U test "
                    "after outlier removal."
                ),
            },
            {
                "item": "Distance label matching",
                "detail": (
                    "Case-sensitive"
                    if case_sensitive_dist
                    else "Case-insensitive: distance labels are uppercased before matching."
                ),
            },
            {
                "item": "FDR adjustment",
                "detail": (
                    "Calculated with Benjamini-Hochberg method."
                    if calculate_fdr
                    else "Not calculated."
                ),
            },
        ]
    )


def save_outputs(
    result: pd.DataFrame,
    skipped: pd.DataFrame,
    method_note: pd.DataFrame,
    output_path: Path,
    simple_columns: List[str],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    suffix = output_path.suffix.lower()

    if suffix in [".xlsx", ".xlsm"]:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            result.to_excel(writer, sheet_name="full_result", index=False)
            result[simple_columns].to_excel(writer, sheet_name="simple_result", index=False)
            skipped.to_excel(writer, sheet_name="skipped_keys", index=False)
            method_note.to_excel(writer, sheet_name="method_note", index=False)

        try:
            format_excel_file(output_path)
        except Exception:
            # Formatting은 optional. 저장된 데이터에는 영향 없음.
            pass

    elif suffix == ".csv":
        result.to_csv(output_path, index=False, encoding="utf-8-sig")

        simple_path = output_path.with_name(f"{output_path.stem}_simple.csv")
        skipped_path = output_path.with_name(f"{output_path.stem}_skipped_keys.csv")
        note_path = output_path.with_name(f"{output_path.stem}_method_note.csv")

        result[simple_columns].to_csv(simple_path, index=False, encoding="utf-8-sig")
        skipped.to_csv(skipped_path, index=False, encoding="utf-8-sig")
        method_note.to_csv(note_path, index=False, encoding="utf-8-sig")

    else:
        raise ValueError(
            f"Unsupported output extension: {suffix}. Use .xlsx, .xlsm, or .csv."
        )


def format_excel_file(output_path: Path) -> None:
    """
    Excel 결과 파일 가독성 개선:
    - 첫 행 freeze
    - filter 적용
    - column width 자동 조정
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(output_path)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 0

            for cell in column_cells:
                if cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))

            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    workbook.save(output_path)


def main() -> None:
    args = parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    try:
        df = read_input_table(input_path)

        result, skipped = build_result_tables(
            df=df,
            step_col=args.step_col,
            item_col=args.item_col,
            value_col=args.value_col,
            dist_col=args.dist_col,
            group_col=args.group_col,
            ref_dist=args.ref_dist,
            blank_group_name=args.blank_group_name,
            min_n_outlier=args.min_n_outlier,
            case_sensitive_dist=args.case_sensitive_dist,
            calculate_fdr=not args.no_fdr,
        )

        simple_columns = [
            args.step_col,
            args.item_col,
            args.group_col,
            "ref_group",
            "comp_group",
            "ref_median",
            "comp_median",
            "p_value",
        ]

        method_note = build_method_note(
            step_col=args.step_col,
            item_col=args.item_col,
            value_col=args.value_col,
            dist_col=args.dist_col,
            group_col=args.group_col,
            ref_dist=args.ref_dist,
            blank_group_name=args.blank_group_name,
            min_n_outlier=args.min_n_outlier,
            case_sensitive_dist=args.case_sensitive_dist,
            calculate_fdr=not args.no_fdr,
        )

        save_outputs(
            result=result,
            skipped=skipped,
            method_note=method_note,
            output_path=output_path,
            simple_columns=simple_columns,
        )

        print(f"[Done] Result saved to: {output_path}")
        print(f"[Done] Number of tests: {len(result):,}")
        print(f"[Done] Number of skipped keys: {len(skipped):,}")

        if len(result) == 0:
            print(
                "[Warning] No test was performed. Check whether E4 exists within each "
                "step_seq/item_id/Group key and whether value is numeric."
            )

    except Exception as exc:
        print(f"[Error] {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()